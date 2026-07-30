# =============================================================================
# Bank Management System - Utilities Module
# =============================================================================
# Author: Bank Management System Team
# Version: 1.0.0
# Description: Shared utility functions for validation, formatting, security,
#              PDF/Excel/CSV export, and UI helpers.
# =============================================================================

import re
import os
import csv
import math
import random
import string
import hashlib
import logging
import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List, Optional, Tuple, Union
from pathlib import Path

import bcrypt

from config import (
    REGEX_EMAIL, REGEX_PHONE, REGEX_PAN, REGEX_AADHAR, REGEX_IFSC,
    REGEX_PINCODE, REGEX_ACCOUNT_NUMBER, REGEX_UPI_ID, REGEX_CARD_NUMBER,
    REGEX_CVV, DATE_FORMAT, DATETIME_FORMAT, DB_DATE_FORMAT,
    DB_DATETIME_FORMAT, EXPORTS_DIR, STATEMENTS_DIR, REPORTS_DIR,
    PIN_LENGTH, PASSWORD_MIN_LENGTH, PASSWORD_REQUIRES_UPPERCASE,
    PASSWORD_REQUIRES_LOWERCASE, PASSWORD_REQUIRES_DIGIT,
    PASSWORD_REQUIRES_SPECIAL, SPECIAL_CHARACTERS, BCRYPT_ROUNDS,
    BANK_NAME, BANK_CODE, FONT_FAMILY_PRIMARY, COLOR_BG_SECONDARY,
    COLOR_ACCENT_PRIMARY, COLOR_TEXT_PRIMARY, COLOR_ACCENT_SUCCESS
)

logger = logging.getLogger(__name__)


# =============================================================================
# SECURITY & HASHING
# =============================================================================

def hash_password(plain_password: str) -> str:
    """
    Hash a password using bcrypt with configured cost factor.

    Args:
        plain_password: Plain text password.

    Returns:
        Hashed password string (includes salt).
    """
    if not plain_password:
        raise ValueError("Password cannot be empty.")
    password_bytes = plain_password.encode("utf-8")
    salt = bcrypt.gensalt(rounds=BCRYPT_ROUNDS)
    hashed = bcrypt.hashpw(password_bytes, salt)
    return hashed.decode("utf-8")


def verify_password(plain_password: str, hashed_password: str) -> bool:
    """
    Verify a plain text password against a bcrypt hash.

    Args:
        plain_password: Plain text password to verify.
        hashed_password: Previously hashed password.

    Returns:
        True if password matches, False otherwise.
    """
    try:
        return bcrypt.checkpw(
            plain_password.encode("utf-8"),
            hashed_password.encode("utf-8")
        )
    except Exception as e:
        logger.error(f"Password verification error: {e}")
        return False


def hash_pin(pin: str) -> str:
    """Hash a numeric PIN using bcrypt."""
    return hash_password(pin)


def verify_pin(pin: str, hashed_pin: str) -> bool:
    """Verify a numeric PIN against its hash."""
    return verify_password(pin, hashed_pin)


def generate_otp(length: int = 6) -> str:
    """Generate a numeric OTP."""
    return "".join(random.choices(string.digits, k=length))


def generate_token(length: int = 32) -> str:
    """Generate a cryptographically secure random token."""
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


# =============================================================================
# INPUT VALIDATION
# =============================================================================

class ValidationResult:
    """Holds the result of a validation check."""

    def __init__(self, is_valid: bool, error_message: str = ""):
        self.is_valid = is_valid
        self.error_message = error_message

    def __bool__(self):
        return self.is_valid

    def __str__(self):
        return self.error_message if not self.is_valid else "Valid"


def validate_email(email: str) -> ValidationResult:
    """Validate email address format."""
    if not email or not email.strip():
        return ValidationResult(False, "Email address is required.")
    if not re.match(REGEX_EMAIL, email.strip()):
        return ValidationResult(False, "Invalid email address format.")
    return ValidationResult(True)


def validate_phone(phone: str) -> ValidationResult:
    """Validate Indian mobile number (10 digits, starts with 6-9)."""
    if not phone or not phone.strip():
        return ValidationResult(False, "Phone number is required.")
    clean_phone = phone.strip().replace(" ", "").replace("-", "")
    if not re.match(REGEX_PHONE, clean_phone):
        return ValidationResult(
            False,
            "Invalid phone number. Must be 10 digits starting with 6-9."
        )
    return ValidationResult(True)


def validate_pan(pan: str) -> ValidationResult:
    """Validate PAN card number format."""
    if not pan or not pan.strip():
        return ValidationResult(False, "PAN number is required.")
    if not re.match(REGEX_PAN, pan.strip().upper()):
        return ValidationResult(
            False,
            "Invalid PAN format. Example: ABCDE1234F"
        )
    return ValidationResult(True)


def validate_aadhar(aadhar: str) -> ValidationResult:
    """Validate Aadhar card number (12 digits)."""
    if not aadhar or not aadhar.strip():
        return ValidationResult(False, "Aadhar number is required.")
    clean = aadhar.strip().replace(" ", "")
    if not re.match(REGEX_AADHAR, clean):
        return ValidationResult(False, "Invalid Aadhar number. Must be 12 digits.")
    return ValidationResult(True)


def validate_ifsc(ifsc: str) -> ValidationResult:
    """Validate IFSC code format."""
    if not ifsc or not ifsc.strip():
        return ValidationResult(False, "IFSC code is required.")
    if not re.match(REGEX_IFSC, ifsc.strip().upper()):
        return ValidationResult(
            False,
            "Invalid IFSC code. Example: SBIN0001234"
        )
    return ValidationResult(True)


def validate_pincode(pincode: str) -> ValidationResult:
    """Validate 6-digit Indian postal pincode."""
    if not pincode or not pincode.strip():
        return ValidationResult(False, "Pincode is required.")
    if not re.match(REGEX_PINCODE, pincode.strip()):
        return ValidationResult(False, "Invalid pincode. Must be 6 digits.")
    return ValidationResult(True)


def validate_password(password: str) -> ValidationResult:
    """
    Validate password against security policy.
    Requires minimum length, uppercase, lowercase, digit, and special character.
    """
    if not password:
        return ValidationResult(False, "Password is required.")
    if len(password) < PASSWORD_MIN_LENGTH:
        return ValidationResult(
            False,
            f"Password must be at least {PASSWORD_MIN_LENGTH} characters long."
        )
    if PASSWORD_REQUIRES_UPPERCASE and not re.search(r'[A-Z]', password):
        return ValidationResult(False, "Password must contain at least one uppercase letter.")
    if PASSWORD_REQUIRES_LOWERCASE and not re.search(r'[a-z]', password):
        return ValidationResult(False, "Password must contain at least one lowercase letter.")
    if PASSWORD_REQUIRES_DIGIT and not re.search(r'\d', password):
        return ValidationResult(False, "Password must contain at least one digit.")
    if PASSWORD_REQUIRES_SPECIAL and not re.search(
        f"[{re.escape(SPECIAL_CHARACTERS)}]", password
    ):
        return ValidationResult(
            False,
            f"Password must contain at least one special character: {SPECIAL_CHARACTERS[:15]}..."
        )
    return ValidationResult(True)


def validate_pin(pin: str) -> ValidationResult:
    """Validate numeric PIN of configured length."""
    if not pin:
        return ValidationResult(False, "PIN is required.")
    if not pin.isdigit():
        return ValidationResult(False, "PIN must contain only digits.")
    if len(pin) != PIN_LENGTH:
        return ValidationResult(False, f"PIN must be exactly {PIN_LENGTH} digits.")
    if len(set(pin)) == 1:
        return ValidationResult(False, "PIN cannot have all identical digits.")
    if pin in ("1234", "4321", "0000", "1111", "9999"):
        return ValidationResult(False, "PIN is too simple. Choose a more secure PIN.")
    return ValidationResult(True)


def validate_amount(amount_str: str, min_amount: float = 1.0) -> ValidationResult:
    """Validate a monetary amount string."""
    if not amount_str or not str(amount_str).strip():
        return ValidationResult(False, "Amount is required.")
    try:
        amount = float(str(amount_str).replace(",", ""))
        if amount <= 0:
            return ValidationResult(False, "Amount must be greater than zero.")
        if amount < min_amount:
            return ValidationResult(
                False, f"Minimum amount is ₹{format_currency(min_amount)}."
            )
        if amount > 10_000_000:  # 1 crore limit
            return ValidationResult(False, "Amount exceeds maximum allowed limit.")
        # Check for more than 2 decimal places
        if round(amount, 2) != amount:
            return ValidationResult(
                False, "Amount cannot have more than 2 decimal places."
            )
        return ValidationResult(True)
    except (ValueError, TypeError):
        return ValidationResult(False, "Invalid amount. Please enter a valid number.")


def validate_date(date_str: str, format_str: str = DATE_FORMAT) -> ValidationResult:
    """Validate a date string against a given format."""
    if not date_str or not date_str.strip():
        return ValidationResult(False, "Date is required.")
    try:
        datetime.datetime.strptime(date_str.strip(), format_str)
        return ValidationResult(True)
    except ValueError:
        return ValidationResult(
            False, f"Invalid date format. Expected: {format_str}"
        )


def validate_dob(dob_str: str) -> ValidationResult:
    """Validate date of birth: must be past date and age 18-100."""
    result = validate_date(dob_str)
    if not result:
        return result
    dob = datetime.datetime.strptime(dob_str, DATE_FORMAT).date()
    today = datetime.date.today()
    age = (today - dob).days // 365
    if dob >= today:
        return ValidationResult(False, "Date of birth must be in the past.")
    if age < 18:
        return ValidationResult(False, "Customer must be at least 18 years old.")
    if age > 100:
        return ValidationResult(False, "Invalid date of birth.")
    return ValidationResult(True)


def validate_required(value: str, field_name: str) -> ValidationResult:
    """Validate that a required field is not empty."""
    if not value or not str(value).strip():
        return ValidationResult(False, f"{field_name} is required.")
    return ValidationResult(True)


def validate_name(name: str, field_name: str = "Name") -> ValidationResult:
    """Validate a person's name (2-100 chars, letters and spaces only)."""
    if not name or not name.strip():
        return ValidationResult(False, f"{field_name} is required.")
    clean = name.strip()
    if len(clean) < 2:
        return ValidationResult(False, f"{field_name} must be at least 2 characters.")
    if len(clean) > 100:
        return ValidationResult(False, f"{field_name} cannot exceed 100 characters.")
    if not re.match(r"^[A-Za-z\s\.\-']+$", clean):
        return ValidationResult(
            False, f"{field_name} can only contain letters, spaces, hyphens, and apostrophes."
        )
    return ValidationResult(True)


# =============================================================================
# FORMATTERS
# =============================================================================

def format_currency(amount: Union[float, Decimal, None], symbol: str = "₹") -> str:
    """
    Format amount as Indian currency with commas.
    Example: 1234567.50 → ₹12,34,567.50
    """
    if amount is None:
        return f"{symbol}0.00"
    try:
        amount = Decimal(str(amount)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        # Indian number format
        num_str = f"{abs(amount):,.2f}"
        # Convert standard comma format to Indian format
        parts = str(abs(amount)).split(".")
        integer_part = parts[0]
        decimal_part = parts[1] if len(parts) > 1 else "00"
        decimal_part = decimal_part[:2].ljust(2, "0")

        # Apply Indian grouping: last 3 digits, then groups of 2
        if len(integer_part) > 3:
            last_three = integer_part[-3:]
            rest = integer_part[:-3]
            groups = []
            while len(rest) > 2:
                groups.append(rest[-2:])
                rest = rest[:-2]
            if rest:
                groups.append(rest)
            groups.reverse()
            formatted_int = ",".join(groups) + "," + last_three
        else:
            formatted_int = integer_part

        sign = "-" if amount < 0 else ""
        return f"{sign}{symbol}{formatted_int}.{decimal_part}"
    except Exception:
        return f"{symbol}{amount}"


def format_account_number(account_no: str) -> str:
    """Format account number: show only last 4 digits, mask rest."""
    if not account_no:
        return "N/A"
    acc = str(account_no).strip()
    if len(acc) <= 4:
        return acc
    return f"{'X' * (len(acc) - 4)}{acc[-4:]}"


def format_card_number(card_no: str) -> str:
    """Format card number: 4-4-4-4 with masking of middle digits."""
    if not card_no:
        return "N/A"
    card = str(card_no).strip().replace(" ", "")
    if len(card) == 16:
        return f"{card[:4]} XXXX XXXX {card[12:]}"
    return card


def format_phone(phone: str) -> str:
    """Format phone number: +91 XXXXX XXXXX."""
    if not phone:
        return "N/A"
    clean = str(phone).strip().replace(" ", "").replace("-", "")
    if len(clean) == 10:
        return f"+91 {clean[:5]} {clean[5:]}"
    return phone


def format_date(date_obj: Union[datetime.date, str, None],
                output_format: str = DATE_FORMAT) -> str:
    """Format a date object or string to display format."""
    if date_obj is None:
        return "N/A"
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.datetime.strptime(date_obj, DB_DATE_FORMAT).date()
        except Exception:
            return str(date_obj)
    return date_obj.strftime(output_format)


def format_datetime(dt_obj: Union[datetime.datetime, str, None]) -> str:
    """Format a datetime object to display format."""
    if dt_obj is None:
        return "N/A"
    if isinstance(dt_obj, str):
        try:
            dt_obj = datetime.datetime.strptime(dt_obj, DB_DATETIME_FORMAT)
        except Exception:
            return str(dt_obj)
    return dt_obj.strftime(DATETIME_FORMAT)


def calculate_age(dob: Union[datetime.date, str]) -> int:
    """Calculate age in years from date of birth."""
    if isinstance(dob, str):
        try:
            dob = datetime.datetime.strptime(dob, DB_DATE_FORMAT).date()
        except Exception:
            return 0
    today = datetime.date.today()
    return (today - dob).days // 365


def calculate_loan_emi(
    principal: float,
    annual_rate: float,
    tenure_months: int
) -> float:
    """
    Calculate monthly EMI using standard amortization formula.

    EMI = P × [r(1+r)^n] / [(1+r)^n - 1]
    where r = monthly rate, n = tenure in months.
    """
    if annual_rate == 0:
        return round(principal / tenure_months, 2)
    monthly_rate = annual_rate / (12 * 100)
    emi = principal * (monthly_rate * (1 + monthly_rate) ** tenure_months) / \
          ((1 + monthly_rate) ** tenure_months - 1)
    return round(emi, 2)


def calculate_fd_maturity(
    principal: float,
    annual_rate: float,
    tenure_months: int,
    compounding: str = "Quarterly"
) -> Tuple[float, float]:
    """
    Calculate FD maturity amount and interest earned.

    Args:
        principal: Initial investment amount.
        annual_rate: Annual interest rate (%).
        tenure_months: FD tenure in months.
        compounding: "Monthly", "Quarterly", "Half-Yearly", "Yearly".

    Returns:
        Tuple of (maturity_amount, interest_earned).
    """
    freq_map = {"Monthly": 12, "Quarterly": 4, "Half-Yearly": 2, "Yearly": 1}
    n = freq_map.get(compounding, 4)
    r = annual_rate / 100
    t = tenure_months / 12

    maturity = principal * (1 + r / n) ** (n * t)
    interest = maturity - principal
    return round(maturity, 2), round(interest, 2)


# =============================================================================
# NUMBER / ID GENERATORS
# =============================================================================

def generate_account_number() -> str:
    """Generate a unique 16-digit account number."""
    # Format: BANK_CODE (4) + BRANCH (3) + RANDOM (9)
    return "".join([
        BANK_CODE[:4].ljust(4, "0"),
        str(random.randint(100, 999)),
        str(random.randint(100000000, 999999999))
    ])


def generate_card_number(network: str = "Visa") -> str:
    """Generate a network-valid 16-digit card number using Luhn algorithm."""
    prefixes = {"Visa": "4", "Mastercard": "5", "RuPay": "6"}
    prefix = prefixes.get(network, "4")

    digits = [int(prefix)]
    while len(digits) < 15:
        digits.append(random.randint(0, 9))

    # Luhn check digit
    total = 0
    for i, d in enumerate(reversed(digits)):
        if i % 2 == 0:
            d *= 2
            if d > 9:
                d -= 9
        total += d
    check_digit = (10 - (total % 10)) % 10
    digits.append(check_digit)

    return "".join(str(d) for d in digits)


def generate_cvv() -> str:
    """Generate a 3-digit CVV."""
    return str(random.randint(100, 999))


def generate_upi_id(name: str, account_suffix: str = "") -> str:
    """Generate a UPI ID in the format name@bank."""
    clean_name = re.sub(r'[^a-zA-Z0-9]', '', name.lower())[:10]
    suffix = account_suffix[-4:] if account_suffix else str(random.randint(1000, 9999))
    return f"{clean_name}{suffix}@apxb"


def generate_reference_number() -> str:
    """Generate a unique transaction reference number."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    random_suffix = "".join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"TXN{timestamp}{random_suffix}"


def generate_loan_account_number() -> str:
    """Generate a loan account number."""
    return f"LN{datetime.datetime.now().strftime('%Y%m')}{random.randint(100000, 999999)}"


def generate_fd_number() -> str:
    """Generate a fixed deposit number."""
    return f"FD{datetime.datetime.now().strftime('%Y%m')}{random.randint(10000, 99999)}"


# =============================================================================
# EXPORT UTILITIES
# =============================================================================

def export_to_csv(
    data: List[Dict],
    filename: str,
    directory: Path = EXPORTS_DIR
) -> str:
    """
    Export data to a CSV file.

    Args:
        data: List of dictionaries to export.
        filename: Output filename (without extension).
        directory: Target directory path.

    Returns:
        Full path of the created CSV file.
    """
    if not data:
        raise ValueError("No data to export.")

    filepath = directory / f"{filename}.csv"
    directory.mkdir(parents=True, exist_ok=True)

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

    logger.info(f"CSV exported: {filepath}")
    return str(filepath)


def export_to_excel(
    data: List[Dict],
    filename: str,
    sheet_name: str = "Data",
    directory: Path = EXPORTS_DIR
) -> str:
    """
    Export data to an Excel (.xlsx) file with formatting.

    Args:
        data: List of dictionaries to export.
        filename: Output filename (without extension).
        sheet_name: Name of the Excel sheet.
        directory: Target directory path.

    Returns:
        Full path of the created Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    if not data:
        raise ValueError("No data to export.")

    directory.mkdir(parents=True, exist_ok=True)
    filepath = directory / f"{filename}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Write headers
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows
    alt_fill = PatternFill("solid", fgColor="E8F0FE")
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = row_data.get(key)
            if isinstance(value, datetime.datetime):
                value = value.strftime(DATETIME_FORMAT)
            elif isinstance(value, datetime.date):
                value = value.strftime(DATE_FORMAT)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(data) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(filepath)
    logger.info(f"Excel exported: {filepath}")
    return str(filepath)


def export_to_pdf(
    data: List[Dict],
    filename: str,
    title: str = "Report",
    directory: Path = EXPORTS_DIR
) -> str:
    """
    Export data to a formatted PDF report.

    Args:
        data: List of dictionaries.
        filename: Output filename (without extension).
        title: Report title displayed at top.
        directory: Target directory.

    Returns:
        Full path of the created PDF file.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Run: pip install reportlab")

    if not data:
        raise ValueError("No data to export.")

    directory.mkdir(parents=True, exist_ok=True)
    filepath = directory / f"{filename}.pdf"

    # Use landscape for wide tables
    page_size = landscape(A4) if len(data[0]) > 6 else A4
    doc = SimpleDocTemplate(str(filepath), pagesize=page_size,
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor("#1F3864"),
        alignment=TA_CENTER, fontName="Helvetica-Bold"
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10, textColor=colors.grey,
        alignment=TA_CENTER
    )

    elements = []

    # Title and metadata
    elements.append(Paragraph(BANK_NAME, title_style))
    elements.append(Paragraph(title, ParagraphStyle(
        "ReportTitle", parent=styles["Heading2"],
        alignment=TA_CENTER, fontSize=14,
        textColor=colors.HexColor("#4f8ef7")
    )))
    elements.append(Paragraph(
        f"Generated on: {datetime.datetime.now().strftime(DATETIME_FORMAT)}",
        subtitle_style
    ))
    elements.append(HRFlowable(width="100%", thickness=1,
                               color=colors.HexColor("#4f8ef7"), spaceAfter=0.3*cm))
    elements.append(Spacer(1, 0.3*cm))

    # Build table
    headers = [k.replace("_", " ").title() for k in data[0].keys()]
    keys = list(data[0].keys())
    table_data = [headers]

    for row in data:
        row_values = []
        for k in keys:
            val = row.get(k)
            if isinstance(val, datetime.datetime):
                val = val.strftime(DATETIME_FORMAT)
            elif isinstance(val, datetime.date):
                val = val.strftime(DATE_FORMAT)
            elif val is None:
                val = "—"
            row_values.append(str(val))
        table_data.append(row_values)

    col_count = len(headers)
    available_width = page_size[0] - 3 * cm
    col_width = available_width / col_count

    table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Data rows
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#EEF2FF")]),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(
        f"Total Records: {len(data)}", subtitle_style
    ))

    doc.build(elements)
    logger.info(f"PDF exported: {filepath}")
    return str(filepath)


# =============================================================================
# UI HELPER UTILITIES
# =============================================================================

def truncate_text(text: str, max_length: int = 30, suffix: str = "...") -> str:
    """Truncate text to max_length and append suffix if needed."""
    if not text:
        return ""
    text = str(text)
    if len(text) <= max_length:
        return text
    return text[:max_length - len(suffix)] + suffix


def mask_sensitive(value: str, visible_chars: int = 4) -> str:
    """Mask sensitive data, showing only last N characters."""
    if not value:
        return "N/A"
    value = str(value).strip()
    if len(value) <= visible_chars:
        return "*" * len(value)
    return "*" * (len(value) - visible_chars) + value[-visible_chars:]


def get_initials(name: str) -> str:
    """Extract initials from a full name (max 2 chars)."""
    if not name:
        return "?"
    parts = name.strip().split()
    if len(parts) == 1:
        return parts[0][0].upper()
    return (parts[0][0] + parts[-1][0]).upper()


def format_bytes(size_bytes: int) -> str:
    """Format byte count as human-readable string."""
    if size_bytes == 0:
        return "0 B"
    size_name = ("B", "KB", "MB", "GB", "TB")
    i = int(math.floor(math.log(size_bytes, 1024)))
    p = math.pow(1024, i)
    s = round(size_bytes / p, 2)
    return f"{s} {size_name[i]}"


def get_greeting() -> str:
    """Return time-appropriate greeting."""
    hour = datetime.datetime.now().hour
    if hour < 12:
        return "Good Morning"
    elif hour < 17:
        return "Good Afternoon"
    else:
        return "Good Evening"


def days_until(target_date: datetime.date) -> int:
    """Calculate number of days until a target date."""
    today = datetime.date.today()
    delta = target_date - today
    return delta.days


def parse_db_date(value: Any) -> Optional[datetime.date]:
    """Safely parse a database date value to Python date object."""
    if value is None:
        return None
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        for fmt in (DB_DATE_FORMAT, DATE_FORMAT, "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def parse_db_datetime(value: Any) -> Optional[datetime.datetime]:
    """Safely parse a database datetime value to Python datetime object."""
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, str):
        for fmt in (DB_DATETIME_FORMAT, DATETIME_FORMAT):
            try:
                return datetime.datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def open_file(filepath: str) -> None:
    """Open a file with the system's default application."""
    import subprocess
    import sys

    try:
        if sys.platform.startswith("win"):
            os.startfile(filepath)
        elif sys.platform.startswith("darwin"):
            subprocess.run(["open", filepath], check=True)
        else:
            subprocess.run(["xdg-open", filepath], check=True)
        logger.info(f"Opened file: {filepath}")
    except Exception as e:
        logger.error(f"Failed to open file {filepath}: {e}")


# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging() -> None:
    """
    Configure application-wide logging with file rotation and console output.
    Call once at application startup.
    """
    import logging.handlers
    from config import (
        LOG_FILE, LOG_FORMAT, LOG_DATE_FORMAT, LOG_LEVEL,
        LOG_MAX_BYTES, LOG_BACKUP_COUNT
    )

    root_logger = logging.getLogger()
    root_logger.setLevel(LOG_LEVEL)

    # Rotating file handler
    file_handler = logging.handlers.RotatingFileHandler(
        LOG_FILE,
        maxBytes=LOG_MAX_BYTES,
        backupCount=LOG_BACKUP_COUNT,
        encoding="utf-8"
    )
    file_handler.setLevel(LOG_LEVEL)
    file_handler.setFormatter(logging.Formatter(LOG_FORMAT, LOG_DATE_FORMAT))

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(message)s",
        "%H:%M:%S"
    ))

    root_logger.addHandler(file_handler)
    root_logger.addHandler(console_handler)

    logging.getLogger("mysql.connector").setLevel(logging.WARNING)
    logging.getLogger("PIL").setLevel(logging.WARNING)
